from django.contrib.auth import get_user_model
from django.db import models
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from rest_framework import generics, viewsets, status, exceptions
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from .models import (
    Project,
    Phase,
    Task,
    TaskImage,
    Attendance,
    AttendanceEntry,
    MaterialRequest,
    RequirementSheet,
    Bid,
    Notification,
    DailyReport,
    ActivityLog,
    SitePhoto,
    ConstructionWorker,
    DailySheetTemplate,
    DailySheetEntry,
    DailySheetCellData,
)
from .permissions import (
    IsAdmin,
    IsProjectManager,
    IsAdminOrProjectManager,
    IsSupervisor,
    IsContractor,
    IsOwner,
    TaskPermission,
    AttendancePermission,
    MaterialRequestPermission,
    BidPermission,
    RequirementSheetPermission,
    SitePhotoPermission,
    ConstructionWorkerPermission,
    DailySheetPermission,
)
from .serializers import (
    UserSerializer,
    ProjectSerializer,
    PhaseSerializer,
    TaskSerializer,
    TaskImageSerializer,
    AttendanceSerializer,
    AttendanceWithEntriesSerializer,
    AttendanceCreateWithEntriesSerializer,
    MaterialRequestSerializer,
    RequirementSheetSerializer,
    BidSerializer,
    NotificationSerializer,
    DailyReportSerializer,
    ActivityLogSerializer,
    SitePhotoSerializer,
    ConstructionWorkerSerializer,
    DailySheetTemplateSerializer,
    DailySheetEntrySerializer,
    DailySheetEntryCreateSerializer,
)

User = get_user_model()



class UserViewSet(viewsets.ModelViewSet):
    """
    Admin-only user management. Create users (any role); delete users (except self).
    Created users can log in at their role's login page.
    PMs can list users to assign supervisors to tasks.
    """
    serializer_class = UserSerializer
    permission_classes = [IsAdmin | IsProjectManager]

    def get_queryset(self):
        """Safe queryset with error handling"""
        try:
            return User.objects.all().order_by('id')
        except Exception as e:
            import logging
            logging.error(f"UserViewSet.get_queryset error: {e}")
            return User.objects.none()

    def get_permissions(self):
        # Only admin can create, update, or delete users
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAdmin()]
        # PM and Admin can list users (for supervisor selection)
        return [IsAdminOrProjectManager()]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.pk == request.user.pk:
            return Response(
                {"detail": "You cannot delete your own account."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ProtectedError:
            return Response(
                {"detail": "Cannot delete user: they have related data (projects, attendance, etc.). Remove or reassign those first."},
                status=status.HTTP_400_BAD_REQUEST,
            )


class RoleAwareTokenSerializer(TokenObtainPairSerializer):
    """
    Extend the default JWT serializer to inject role information.
    This lets the frontend immediately know which dashboard to show.
    """

    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token["role"] = getattr(user, "role", None)
        return token

    def validate(self, attrs):
        data = super().validate(attrs)
        data["user"] = UserSerializer(self.user).data
        return data


class RoleAwareTokenView(TokenObtainPairView):
    serializer_class = RoleAwareTokenSerializer
    permission_classes = [AllowAny]


class ContractorRegistrationView(generics.CreateAPIView):
    """
    External contractor self-registration.
    Creates a User with role='contractor'. Admin can later assign bids/projects.
    """

    permission_classes = [AllowAny]
    serializer_class = UserSerializer

    def perform_create(self, serializer):
        user = serializer.save(role=User.Role.CONTRACTOR)
        # Ensure external contractors cannot access Django admin unless promoted.
        user.is_staff = False
        user.is_superuser = False
        user.save(update_fields=["is_staff", "is_superuser"])


class CurrentUserView(generics.RetrieveAPIView):
    """
    Simple endpoint for the frontend to fetch the authenticated user's data.
    """

    serializer_class = UserSerializer

    def get_object(self):
        return self.request.user


class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer
    permission_classes = [IsAdmin | IsProjectManager | IsOwner | IsSupervisor]

    def get_queryset(self):
        user = self.request.user
        try:
            queryset = Project.objects.select_related("project_manager").prefetch_related("owners", "supervisors")
            
            if user.role in ["admin", "owner"]:
                return queryset.all()
            if user.role == "project_manager":
                return queryset.filter(project_manager=user)
            if user.role == "supervisor":
                # Supervisors should see projects explicitly assigned to them.
                return queryset.filter(supervisors=user)
            return Project.objects.none()
        except Exception as e:
            # Log error but return empty queryset instead of crashing
            import logging
            logging.error(f"ProjectViewSet.get_queryset error: {e}")
            return Project.objects.none()

    def get_permissions(self):
        # Admin can create/update/delete.
        # PM might need to update progress? No, "Track and update project progress percentage".
        # So PM can update.
        if self.action in ["create", "destroy"]:
            return [IsAdmin()]
        if self.action in ["update", "partial_update"]:
            return [IsAdmin() | IsProjectManager()]
        return super().get_permissions()

    def perform_create(self, serializer):
        # Business rule: Admin must assign a project manager at creation time.
        if serializer.validated_data.get("project_manager") is None:
            raise exceptions.ValidationError(
                {"project_manager_id": "Project Manager assignment is required."}
            )
        project = serializer.save(created_by=self.request.user)
        # Notify assigned project manager and site engineers (supervisors)
        project_manager = project.project_manager
        if project_manager:
            Notification.objects.create(
                user=project_manager,
                message=f"You have been assigned as Project Manager for project: {project.name}.",
                related_project=project,
            )
        for supervisor in project.supervisors.all():
            Notification.objects.create(
                user=supervisor,
                message=f"You have been assigned as Site Engineer for project: {project.name}.",
                related_project=project,
            )


    @action(detail=True, methods=["get"], permission_classes=[IsAdmin | IsProjectManager | IsOwner])
    def detail_stats(self, request, pk=None):
        """Get detailed stats for a project including photos and task completion"""
        # Get project directly without queryset filtering to avoid 404
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)
        
        # Check if user has permission to view this project
        user = request.user
        if user.role == "project_manager" and project.project_manager != user:
            return Response(
                {"detail": "You do not have permission to view this project."},
                status=status.HTTP_403_FORBIDDEN
            )
        # Owners and admins can view all projects (no restriction)
        
        
        # Get photos for this project
        photos = SitePhoto.objects.filter(project=project).order_by("-uploaded_at")
        
        # Get task completion stats grouped by date
        tasks = Task.objects.filter(phase__project=project).select_related("phase")
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status=Task.Status.COMPLETED).count()
        in_progress_tasks = tasks.filter(status=Task.Status.IN_PROGRESS).count()
        pending_tasks = tasks.filter(status=Task.Status.PENDING).count()
        blocked_tasks = tasks.filter(status=Task.Status.BLOCKED).count()
        
        # Get task completion timeline (for graphs)
        from django.db.models import Count, Q
        from django.db.models.functions import TruncDate
        
        task_timeline = (
            tasks.filter(status=Task.Status.COMPLETED)
            .annotate(completion_date=TruncDate("updated_at"))
            .values("completion_date")
            .annotate(count=Count("id"))
            .order_by("completion_date")
        )
        
        return Response({
            "project": ProjectSerializer(project).data,
            "photos": SitePhotoSerializer(photos, many=True).data,
            "task_stats": {
                "total": total_tasks,
                "completed": completed_tasks,
                "in_progress": in_progress_tasks,
                "pending": pending_tasks,
                "blocked": blocked_tasks,
                "completion_percentage": round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2),
            },
            "task_timeline": list(task_timeline),
        })


class PhaseViewSet(viewsets.ModelViewSet):
    serializer_class = PhaseSerializer
    permission_classes = [IsAdmin | IsProjectManager | IsOwner | IsSupervisor]

    def get_queryset(self):
        """Safely get phases with error handling and project filtering"""
        try:
            user = self.request.user
            qs = Phase.objects.select_related('project').prefetch_related('tasks')
            
            # Support filtering by project
            project_id = self.request.query_params.get('project')
            if project_id:
                qs = qs.filter(project_id=project_id)
            
            if user.role in ["admin", "owner"]:
                return qs.all().order_by('project', 'order')
            if user.role == "project_manager":
                return qs.filter(project__project_manager=user).order_by('project', 'order')
            if user.role == "supervisor":
                # Supervisors see phases for projects they're assigned to
                return qs.filter(project__supervisors=user).distinct().order_by('project', 'order')
            return Phase.objects.none()
        except Exception as e:
            import logging
            logging.error(f"PhaseViewSet.get_queryset error: {e}")
            return Phase.objects.none()

    def get_permissions(self):
        # Admin and PM can create/update/delete phases
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsProjectManager() | IsAdmin()]
        return [IsAdmin() | IsProjectManager() | IsOwner() | IsSupervisor()]

    @action(detail=False, methods=["get"], url_path="by-project/(?P<project_id>[^/.]+)")
    def by_project(self, request, project_id=None):
        """Get all phases for a specific project, ordered by sequence"""
        try:
            phases = self.get_queryset().filter(project_id=project_id).order_by('order')
            serializer = self.get_serializer(phases, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {"detail": f"Error fetching phases: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [TaskPermission]

    def get_queryset(self):
        user = self.request.user
        queryset = Task.objects.select_related("supervisor") # Optimize supervisor fetch
        
        if user.role in ["admin", "owner"]:
            return queryset.all()
        if user.role == "project_manager":
            return queryset.filter(phase__project__project_manager=user)
        if user.role == "supervisor":
            return queryset.filter(supervisor=user)
        return Task.objects.none()

    @action(detail=True, methods=["post"], permission_classes=[IsProjectManager | IsAdmin])
    def assign_supervisor(self, request, pk=None):
        """Assign or reassign a task to a supervisor (Site Manager)"""
        task = self.get_object()
        supervisor_id = request.data.get("supervisor_id")
        
        if not supervisor_id:
            return Response(
                {"detail": "supervisor_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            supervisor = User.objects.get(id=supervisor_id, role="supervisor")
        except User.DoesNotExist:
            return Response(
                {"detail": "Supervisor not found or user is not a supervisor"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Update task
        old_supervisor = task.supervisor
        task.supervisor = supervisor
        task.save()
        
        # Create notification for the new supervisor
        Notification.objects.create(
            user=supervisor,
            message=f"You have been assigned to task: {task.title}",
            related_project=task.phase.project,
        )
        
        # Optionally notify old supervisor if there was one
        if old_supervisor and old_supervisor != supervisor:
            Notification.objects.create(
                user=old_supervisor,
                message=f"Task '{task.title}' has been reassigned to another supervisor",
                related_project=task.phase.project,
            )
        
        return Response({
            "detail": "Task assigned successfully",
            "task": TaskSerializer(task).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[IsSupervisor | IsAdmin])
    def accept_task(self, request, pk=None):
        """Site Manager accepts a task"""
        task = self.get_object()
        
        if task.supervisor != request.user and request.user.role != "admin":
            return Response(
                {"detail": "You are not assigned to this task"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if task.approval_status != Task.ApprovalStatus.PENDING_ACCEPTANCE:
            return Response(
                {"detail": f"Task is already {task.approval_status}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task.approval_status = Task.ApprovalStatus.ACCEPTED
        task.status = Task.Status.IN_PROGRESS
        task.save()
        
        # Notify PM
        pm = task.phase.project.project_manager
        if pm:
            Notification.objects.create(
                user=pm,
                message=f"Task '{task.title}' has been accepted by {request.user.username}",
                related_project=task.phase.project,
            )
        
        return Response({
            "detail": "Task accepted successfully",
            "task": TaskSerializer(task).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[IsSupervisor | IsAdmin])
    def reject_task(self, request, pk=None):
        """Site Manager rejects a task"""
        task = self.get_object()
        
        if task.supervisor != request.user and request.user.role != "admin":
            return Response(
                {"detail": "You are not assigned to this task"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if task.approval_status != Task.ApprovalStatus.PENDING_ACCEPTANCE:
            return Response(
                {"detail": f"Task is already {task.approval_status}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task.approval_status = Task.ApprovalStatus.REJECTED
        task.save()
        
        # Notify PM
        pm = task.phase.project.project_manager
        if pm:
            Notification.objects.create(
                user=pm,
                message=f"Task '{task.title}' has been rejected by {request.user.username}",
                related_project=task.phase.project,
            )
        
        return Response({
            "detail": "Task rejected",
            "task": TaskSerializer(task).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[IsSupervisor | IsAdmin])
    def mark_done(self, request, pk=None):
        """Site Manager marks task as done (pending PM approval)"""
        task = self.get_object()
        
        if task.supervisor != request.user and request.user.role != "admin":
            return Response(
                {"detail": "You are not assigned to this task"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if task.approval_status != Task.ApprovalStatus.ACCEPTED:
            return Response(
                {"detail": "Task must be accepted before marking as done"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task.approval_status = Task.ApprovalStatus.PENDING_COMPLETION
        task.progress_percent = 100
        task.save()
        
        # Notify PM
        pm = task.phase.project.project_manager
        if pm:
            Notification.objects.create(
                user=pm,
                message=f"Task '{task.title}' has been marked as done by {request.user.username}. Please review and approve.",
                related_project=task.phase.project,
            )
        
        return Response({
            "detail": "Task marked as done, awaiting PM approval",
            "task": TaskSerializer(task).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[IsProjectManager | IsAdmin])
    def approve_completion(self, request, pk=None):
        """PM approves task completion"""
        task = self.get_object()
        
        if task.approval_status != Task.ApprovalStatus.PENDING_COMPLETION:
            return Response(
                {"detail": "Task is not pending completion approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task.approval_status = Task.ApprovalStatus.APPROVED
        task.status = Task.Status.COMPLETED
        task.save()
        
        # Notify Site Manager
        if task.supervisor:
            Notification.objects.create(
                user=task.supervisor,
                message=f"Task '{task.title}' has been approved and marked as completed",
                related_project=task.phase.project,
            )
        
        return Response({
            "detail": "Task completion approved",
            "task": TaskSerializer(task).data
        }, status=status.HTTP_200_OK)


class TaskImageViewSet(viewsets.ModelViewSet):
    queryset = TaskImage.objects.all()
    serializer_class = TaskImageSerializer
    permission_classes = [TaskPermission] # Similar access rules

    def get_queryset(self):
        user = self.request.user
        if user.role in ["admin", "owner"]:
            return TaskImage.objects.all()
        if user.role == "project_manager":
            return TaskImage.objects.filter(task__phase__project__project_manager=user)
        if user.role == "supervisor":
             # Supervisors see images for their tasks
            return TaskImage.objects.filter(task__supervisor=user)
        return TaskImage.objects.none()

    def perform_create(self, serializer):
        # Only supervisors upload, but we'll enforce logged in user
        serializer.save(uploaded_by=self.request.user)


class AttendanceViewSet(viewsets.ModelViewSet):
    permission_classes = [AttendancePermission]

    def get_serializer_class(self):
        if self.action == "list" or self.action == "retrieve":
            return AttendanceWithEntriesSerializer
        if self.action == "submit_with_entries":
            return AttendanceCreateWithEntriesSerializer
        return AttendanceSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Attendance.objects.prefetch_related("entries__worker")
        if user.role in ["admin", "owner"]:
            return qs.all()
        if user.role == "project_manager":
            return qs.filter(project__project_manager=user)
        if user.role == "supervisor":
            return qs.filter(supervisor=user)
        return qs.none()

    def perform_create(self, serializer):
        serializer.save(supervisor=self.request.user)

    @action(detail=False, methods=["post"], url_path="submit-with-entries")
    def submit_with_entries(self, request):
        """Supervisor submits attendance with per-worker present/absent."""
        ser = AttendanceCreateWithEntriesSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        attendance = ser.save()
        return Response(
            AttendanceWithEntriesSerializer(attendance).data,
            status=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        raise exceptions.PermissionDenied("Attendance records are immutable.")

    def partial_update(self, request, *args, **kwargs):
        raise exceptions.PermissionDenied("Attendance records are immutable.")


class SitePhotoViewSet(viewsets.ModelViewSet):
    serializer_class = SitePhotoSerializer
    permission_classes = [SitePhotoPermission]

    def get_queryset(self):
        user = self.request.user
        qs = SitePhoto.objects.select_related("project", "uploaded_by").order_by("-uploaded_at")
        if user.role in ["admin", "owner"]:
            return qs.all()
        if user.role == "project_manager":
            return qs.filter(project__project_manager=user)
        if user.role == "supervisor":
            return qs.filter(uploaded_by=user)
        return qs.none()

    def perform_create(self, serializer):
        user = self.request.user
        project = serializer.validated_data.get("project")
        if user.role == "supervisor" and project and not project.supervisors.filter(pk=user.pk).exists():
            raise exceptions.PermissionDenied("You can only upload photos for projects you are assigned to.")
        serializer.save(uploaded_by=user)


class ConstructionWorkerViewSet(viewsets.ModelViewSet):
    serializer_class = ConstructionWorkerSerializer
    permission_classes = [ConstructionWorkerPermission]

    def get_queryset(self):
        user = self.request.user
        qs = ConstructionWorker.objects.select_related("project", "added_by")
        project_id = self.request.query_params.get("project")
        if project_id:
            qs = qs.filter(project_id=project_id)
        if user.role in ["admin", "owner"]:
            return qs.all()
        if user.role == "project_manager":
            return qs.filter(project__project_manager=user)
        if user.role == "supervisor":
            return qs.filter(project__supervisors=user)
        return qs.none()

    def perform_create(self, serializer):
        serializer.save(added_by=self.request.user)


class MaterialRequestViewSet(viewsets.ModelViewSet):
    serializer_class = MaterialRequestSerializer
    permission_classes = [MaterialRequestPermission]

    def get_queryset(self):
        user = self.request.user
        
        # Contractors only see PUBLISHED requests
        if user.role == "contractor":
            return MaterialRequest.objects.filter(status=MaterialRequest.Status.PUBLISHED)
        
        if user.role == "project_manager":
            return MaterialRequest.objects.filter(project__project_manager=user)
        
        if user.role == "supervisor":
            return MaterialRequest.objects.filter(raised_by=user)
            
        if user.role in ["admin", "owner"]:
            return MaterialRequest.objects.all()
            
        return MaterialRequest.objects.none()

    def perform_create(self, serializer):
        serializer.save(raised_by=self.request.user)


class RequirementSheetViewSet(viewsets.ModelViewSet):
    """
    Admin creates/publishes requirement sheets (RFQs).
    Contractors can only see published sheets.
    """

    serializer_class = RequirementSheetSerializer
    permission_classes = [RequirementSheetPermission]

    def get_queryset(self):
        user = self.request.user
        if user.role == "contractor":
            return RequirementSheet.objects.filter(
                status=RequirementSheet.Status.PUBLISHED
            ).order_by("-created_at")
        if user.role == "project_manager":
            return RequirementSheet.objects.filter(
                project__project_manager=user
            ).order_by("-created_at")
        if user.role == "owner":
            return RequirementSheet.objects.all().order_by("-created_at")
        if user.role == "admin":
            return RequirementSheet.objects.all().order_by("-created_at")
        return RequirementSheet.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class BidViewSet(viewsets.ModelViewSet):
    serializer_class = BidSerializer
    permission_classes = [BidPermission]

    def get_queryset(self):
        user = self.request.user
        
        # Contractors see only their own bids
        if user.role == "contractor":
            return Bid.objects.filter(contractor=user)
            
        if user.role == "project_manager":
             # PM sees bids for their project's MRs
             return Bid.objects.filter(material_request__project__project_manager=user)
             
        if user.role in ["admin", "owner"]:
            return Bid.objects.all()
            
        return Bid.objects.none()

    def perform_create(self, serializer):
        # Ensure contractor is set
        serializer.save(contractor=self.request.user)


class NotificationViewSet(viewsets.ModelViewSet):
    """
    In-app notifications for the authenticated user.
    Includes helpers to mark items as read/unread in bulk.
    """

    serializer_class = NotificationSerializer

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by(
            "-created_at"
        )

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save(update_fields=["is_read"])
        return Response({"status": "marked_read"})

    @action(detail=True, methods=["post"])
    def mark_unread(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = False
        notification.save(update_fields=["is_read"])
        return Response({"status": "marked_unread"})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        count = (
            Notification.objects.filter(user=request.user, is_read=False)
            .update(is_read=True)
        )
        return Response({"status": "all_marked_read", "updated": count})


class DailyReportViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = DailyReportSerializer
    permission_classes = [IsAdmin | IsProjectManager | IsOwner]

    def get_queryset(self):
        user = self.request.user
        if user.role in ["admin", "owner"]:
            return DailyReport.objects.all()
        if user.role == "project_manager":
            return DailyReport.objects.filter(project__project_manager=user)
        return DailyReport.objects.none()


class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Azure DevOps-style history feed.
    """
    queryset = ActivityLog.objects.select_related("user").all()
    serializer_class = ActivityLogSerializer
    permission_classes = [IsAdmin] # Req: "Access complete activity... -> Admin"

    def get_queryset(self):
        qs = super().get_queryset()
        entity_type = self.request.query_params.get("entity_type")
        entity_id = self.request.query_params.get("entity_id")
        user_id = self.request.query_params.get("user_id")

        if entity_type:
            qs = qs.filter(entity_type=entity_type)
        if entity_id:
            qs = qs.filter(entity_id=entity_id)
        if user_id:
            qs = qs.filter(user_id=user_id)

        return qs


class DashboardViewSet(viewsets.ViewSet):
    """
    Owner dashboard for aggregated analytics.
    """
    permission_classes = [IsOwner | IsAdmin]

    @action(detail=False, methods=["get"])
    def stats(self, request):
        # Simple aggregated stats
        total_projects = Project.objects.count()
        projects_by_city = Project.objects.values("city").annotate(
            count=models.Count("id"),
            avg_progress=models.Avg("progress_percent")
        )
        projects_by_location = Project.objects.values("location").annotate(
             count=models.Count("id")
        )

        return Response({
            "total_projects": total_projects,
            "by_city": projects_by_city,
            "by_location": projects_by_location,
        })


class DailySheetTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing daily sheet templates.
    Project Manager creates templates; all roles can read.
    """
    serializer_class = DailySheetTemplateSerializer
    permission_classes = [DailySheetPermission]

    def get_queryset(self):
        user = self.request.user
        qs = DailySheetTemplate.objects.select_related("project", "created_by")
        
        if user.role in ["admin", "owner"]:
            return qs.all()
        if user.role == "project_manager":
            return qs.filter(project__project_manager=user)
        if user.role == "supervisor":
            # Supervisors see templates for projects they're assigned to
            return qs.filter(project__supervisors=user)
        return DailySheetTemplate.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class DailySheetEntryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing daily sheet entries.
    Supervisors fill entries; PM/Admin/Owner can view.
    """
    permission_classes = [DailySheetPermission]

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return DailySheetEntryCreateSerializer
        return DailySheetEntrySerializer

    def get_queryset(self):
        user = self.request.user
        qs = DailySheetEntry.objects.select_related(
            "template", "project", "filled_by"
        ).prefetch_related("cell_data")
        
        if user.role in ["admin", "owner"]:
            return qs.all()
        if user.role == "project_manager":
            return qs.filter(project__project_manager=user)
        if user.role == "supervisor":
            return qs.filter(filled_by=user)
        return DailySheetEntry.objects.none()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        entry = serializer.save()
        return Response(
            DailySheetEntrySerializer(entry).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=["get"], url_path="download-excel")
    def download_excel(self, request, pk=None):
        """
        Generate and download Excel file for a specific daily sheet entry.
        """
        entry = self.get_object()
        template = entry.template
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Sheet"
        
        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(template.column_headings) + 1)
        title_cell = ws.cell(row=1, column=1, value=f"{template.name} - {entry.date}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        
        # Project info
        ws.cell(row=2, column=1, value=f"Project: {entry.project.name}")
        ws.cell(row=2, column=1).font = Font(bold=True)
        
        # Headers row (row 4)
        header_row = 4
        ws.cell(row=header_row, column=1, value="")  # Top-left corner cell
        for col_idx, col_heading in enumerate(template.column_headings, start=2):
            cell = ws.cell(row=header_row, column=col_idx, value=col_heading)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Data rows
        cell_data_dict = {(cd.row_index, cd.column_index): cd.value for cd in entry.cell_data.all()}
        
        for row_idx, row_heading in enumerate(template.row_headings):
            data_row = header_row + row_idx + 1
            # Row heading
            cell = ws.cell(row=data_row, column=1, value=row_heading)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Cell values
            for col_idx in range(len(template.column_headings)):
                value = cell_data_dict.get((row_idx, col_idx), "")
                ws.cell(row=data_row, column=col_idx + 2, value=value)
        
        # Notes section
        if entry.notes:
            notes_row = header_row + len(template.row_headings) + 2
            ws.cell(row=notes_row, column=1, value="Notes:")
            ws.cell(row=notes_row, column=1).font = Font(bold=True)
            ws.cell(row=notes_row + 1, column=1, value=entry.notes)
        
        # Auto-size columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        filename = f"{template.name}_{entry.date}.xlsx"
        response = HttpResponse(
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="by-project/(?P<project_id>[^/.]+)")
    def by_project(self, request, project_id=None):
        """Get all daily sheet entries for a specific project."""
        entries = self.get_queryset().filter(project_id=project_id).order_by("-date")
        serializer = self.get_serializer(entries, many=True)
        return Response(serializer.data)


